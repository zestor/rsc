#!/usr/bin/env python3
"""Annotate an XLSX with review issues as cell comments.

Usage:
    python annotate_xlsx.py input.xlsx output.xlsx

Reads issues from document_review_state.json in the working directory.
Uses openpyxl Comments attached to cells containing the
original text. Falls back to cell A1 if text not found.
"""

import json
import shutil
import sys
from pathlib import Path

from constants import STATE_FILENAME
from models import format_comment
from openpyxl import load_workbook
from openpyxl.comments import Comment

AUTHOR = "Perplexity"


def load_issues():
    """Load issues from document_review_state.json."""
    path = Path(STATE_FILENAME)
    if not path.exists():
        print(
            f"Error: {STATE_FILENAME} not found.",
            file=sys.stderr,
        )
        sys.exit(1)
    state = json.loads(path.read_text(encoding="utf-8"))
    return list(state.get("issues", {}).values())


def find_cell(ws, text):
    """Find first cell containing the given text."""
    target = str(text).strip().lower()
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            if target in str(cell.value).strip().lower():
                return cell
    return None


def find_worksheet(wb, location):
    """Find a worksheet by name (case-insensitive)."""
    for ws in wb.worksheets:
        if ws.title.lower() == location.lower():
            return ws
    return None


def _place_on_cell(ws, anchor, comment):
    """Try to place a comment directly on the given cell reference."""
    try:
        cell = ws[anchor]
        cell.comment = comment
        return True
    except (KeyError, ValueError):
        return False


def annotate(input_path, output_path):
    """Add comment annotations to the XLSX."""
    issues = load_issues()
    if not issues:
        print("No issues to annotate.")
        return

    shutil.copy2(input_path, output_path)
    wb = load_workbook(output_path)
    annotated = 0

    for issue in issues:
        text = format_comment(issue)
        comment = Comment(text, AUTHOR)
        original = issue.get("original_text", "")
        location = issue.get("location", "")
        anchor = issue.get("anchor")

        placed = False

        target_ws = find_worksheet(wb, location) if location else None

        if target_ws and anchor:
            placed = _place_on_cell(target_ws, anchor, comment)

        if not placed and target_ws:
            cell = find_cell(target_ws, original)
            if cell:
                cell.comment = comment
                placed = True

        if not placed:
            for ws in wb.worksheets:
                cell = find_cell(ws, original)
                if cell:
                    cell.comment = comment
                    placed = True
                    break

        if not placed:
            ws = wb.worksheets[0]
            fallback = ws["A1"]
            if fallback.comment:
                existing = fallback.comment.text
                fallback.comment = Comment(
                    f"{existing}\n\n---\n\n{text}",
                    AUTHOR,
                )
            else:
                fallback.comment = comment

        annotated += 1

    wb.save(output_path)
    print(f"Added {annotated} comments to {output_path}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print(
            "Usage: annotate_xlsx.py <input.xlsx> <output.xlsx>",
            file=sys.stderr,
        )
        sys.exit(1)
    annotate(sys.argv[1], sys.argv[2])
