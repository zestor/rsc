"""
Recalculate all formulas in an Excel file via LibreOffice headless, then
scan for residual errors (#REF!, #DIV/0!, etc.).

Usage:
    python recalc.py <excel_file> [timeout_seconds]
"""

import json
import re
import shutil
import subprocess
import sys
import zipfile
from pathlib import Path

from _soffice import macro_dir, run_soffice, soffice_env
from openpyxl import load_workbook

RECALC_MACRO = """\
<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" \
script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""

MACRO_SCRIPT_URI = "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application"

EXCEL_ERRORS = ("#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A")

TABLE_STYLE_RE = re.compile(rb"<tableStyleInfo[^/>]*/>")

MAX_ERROR_LOCATIONS = 20

DEFAULT_TIMEOUT = 30


def _ensure_macro() -> bool:
    macro_file = macro_dir() / "Module1.xba"

    if macro_file.exists() and "RecalculateAndSave" in macro_file.read_text():
        return True

    if not macro_file.parent.exists():
        subprocess.run(
            ["soffice", "--headless", "--terminate_after_init"],
            capture_output=True,
            timeout=10,
            env=soffice_env(),
        )
        macro_file.parent.mkdir(parents=True, exist_ok=True)

    try:
        macro_file.write_text(RECALC_MACRO)
        return True
    except Exception:
        return False


def _snapshot_table_styles(path: str) -> dict[str, bytes]:
    styles: dict[str, bytes] = {}
    with zipfile.ZipFile(path, "r") as zf:
        for name in zf.namelist():
            if name.startswith("xl/tables/") and name.endswith(".xml"):
                m = TABLE_STYLE_RE.search(zf.read(name))
                if m:
                    styles[name] = m.group(0)
    return styles


def _patch_table_style(data: bytes, style_element: bytes) -> bytes:
    if TABLE_STYLE_RE.search(data):
        return TABLE_STYLE_RE.sub(style_element, data)
    return data.replace(b"</table>", style_element + b"</table>")


def _restore_table_styles(path: str, styles: dict[str, bytes]) -> None:
    if not styles:
        return

    tmp = path + ".tmp"
    try:
        with zipfile.ZipFile(path, "r") as zin, zipfile.ZipFile(tmp, "w") as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename in styles:
                    data = _patch_table_style(data, styles[item.filename])
                zout.writestr(item, data)
        shutil.move(tmp, path)
    except Exception:
        import os  # noqa: PLC0415

        if os.path.exists(tmp):
            os.remove(tmp)


def _scan_errors(path: str) -> dict[str, list[str]]:
    wb = load_workbook(path, data_only=True)
    found: dict[str, list[str]] = {e: [] for e in EXCEL_ERRORS}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    for err in EXCEL_ERRORS:
                        if err in cell.value:
                            found[err].append(f"{ws.title}!{cell.coordinate}")
                            break
    wb.close()
    return found


def _count_formulas(path: str) -> int:
    wb = load_workbook(path, data_only=False)
    count = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    count += 1
    wb.close()
    return count


def recalc(filename: str, timeout: int = DEFAULT_TIMEOUT) -> dict:
    path = Path(filename)
    if not path.exists():
        return {"error": f"File {filename} does not exist"}

    abs_path = str(path.absolute())

    if not _ensure_macro():
        return {"error": "Failed to setup LibreOffice macro"}

    try:
        table_styles = _snapshot_table_styles(abs_path)
    except Exception:
        table_styles = {}

    try:
        result = run_soffice(
            ["--headless", "--norestore", MACRO_SCRIPT_URI, abs_path], timeout=timeout
        )
    except subprocess.TimeoutExpired:
        return {"error": f"LibreOffice timed out after {timeout}s"}

    if result.returncode != 0:
        error_msg = result.stderr or "Unknown error during recalculation"
        if "Module1" in error_msg and "RecalculateAndSave" not in error_msg:
            return {"error": "LibreOffice macro not configured properly"}
        return {"error": error_msg}

    try:
        _restore_table_styles(abs_path, table_styles)
    except Exception:
        pass

    try:
        errors = _scan_errors(abs_path)
        total_errors = sum(len(locs) for locs in errors.values())

        summary = {}
        for err_type, locations in errors.items():
            if locations:
                summary[err_type] = {
                    "count": len(locations),
                    "locations": locations[:MAX_ERROR_LOCATIONS],
                }

        return {
            "status": "success" if total_errors == 0 else "errors_found",
            "total_errors": total_errors,
            "total_formulas": _count_formulas(abs_path),
            "error_summary": summary,
        }
    except Exception as e:
        return {"error": str(e)}


def main() -> None:
    if len(sys.argv) < 2:
        print("Usage: python recalc.py <excel_file> [timeout_seconds]")
        sys.exit(1)

    filename = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_TIMEOUT
    print(json.dumps(recalc(filename, timeout), indent=2))


if __name__ == "__main__":
    main()
